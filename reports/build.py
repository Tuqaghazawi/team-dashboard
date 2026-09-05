"""Building the prep-clinic activity report.

The prep-clinic coordinator and the chairman need the same numbers weekly and
monthly: how many patients were registered, split by specialty, by diagnosis,
by age band, and by the consultant they were assigned to.
"""

import calendar
from collections import Counter
from datetime import timedelta
from io import BytesIO

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

AGE_BANDS = [(0, 39, "under 40"), (40, 54, "40-54"), (55, 69, "55-69"), (70, 200, "70+")]

# The workbook gets downloaded and passed around, so it carries the label too.
NOT_CLINICAL = "NOT FOR CLINICAL USE - prototype on synthetic data"

HEADER_FILL = PatternFill("solid", fgColor="028090")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="13343B")


def period_range(period, today=None):
    """(start, end, label) for 'week' or 'month'."""
    today = today or timezone.localdate()
    if period == "week":
        start = today - timedelta(days=today.weekday())
        end = start + timedelta(days=6)
        return start, end, f"Week of {start:%d %b %Y}"
    start = today.replace(day=1)
    end = today.replace(day=calendar.monthrange(today.year, today.month)[1])
    return start, end, f"{start:%B %Y}"


def age_band(age):
    for low, high, label in AGE_BANDS:
        if low <= age <= high:
            return label
    return "unknown"


def collect(patients, start, end):
    """Patients registered inside the period, plus the counts the report shows."""
    rows = [
        p for p in patients.select_related("team")
        if start <= timezone.localtime(p.registered_at).date() <= end
    ]
    # Ordered (value, count) lists rather than Counters: a Counter is a dict
    # subclass whose __missing__ returns 0, so a Django template asking for
    # ".most_common" gets the integer 0 instead of the method.
    return {
        "patients": rows,
        "total": len(rows),
        "by_specialty": Counter(p.get_specialty_display() for p in rows).most_common(),
        "by_diagnosis": Counter(p.diagnosis for p in rows).most_common(),
        "by_age": Counter(age_band(p.age) for p in rows).most_common(),
        "by_consultant": Counter(p.team.consultant for p in rows).most_common(),
    }


def build_workbook(data, label):
    """An .xlsx workbook: a summary sheet plus the full patient list."""
    wb = Workbook()

    summary = wb.active
    summary.title = "Summary"
    _title(summary, f"Prep clinic report — {label}")
    summary["A3"] = "Total patients registered"
    summary["A3"].font = Font(bold=True)
    summary["B3"] = data["total"]

    row = 5
    for heading, counts in (
        ("By specialty", data["by_specialty"]),
        ("By consultant", data["by_consultant"]),
        ("By age band", data["by_age"]),
        ("By diagnosis", data["by_diagnosis"]),
    ):
        row = _counter_block(summary, row, heading, counts)
    _autofit(summary)

    detail = wb.create_sheet("Patients")
    detail["A1"] = NOT_CLINICAL
    detail["A1"].font = Font(bold=True, color="9A281B", size=11)
    headers = [
        "MRN", "Name", "Date of birth", "Age", "Specialty",
        "Diagnosis", "Consultant", "Stage", "Registered",
    ]
    _header_row(detail, headers, row=2)
    for i, p in enumerate(data["patients"], start=3):
        detail.cell(i, 1, p.mrn)
        detail.cell(i, 2, p.name)
        detail.cell(i, 3, p.date_of_birth.isoformat())
        detail.cell(i, 4, p.age)
        detail.cell(i, 5, p.get_specialty_display())
        detail.cell(i, 6, p.diagnosis)
        detail.cell(i, 7, p.team.consultant)
        detail.cell(i, 8, p.get_stage_display())
        detail.cell(i, 9, timezone.localtime(p.registered_at).date().isoformat())
    detail.freeze_panes = "A3"
    _autofit(detail)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def _title(sheet, text):
    sheet["A1"] = text
    sheet["A1"].font = TITLE_FONT
    sheet["A2"] = NOT_CLINICAL
    sheet["A2"].font = Font(bold=True, color="9A281B", size=11)


def _counter_block(sheet, row, heading, counts):
    sheet.cell(row, 1, heading).font = Font(bold=True, color="028090")
    row += 1
    sheet.cell(row, 1, "Value").font = HEADER_FONT
    sheet.cell(row, 1).fill = HEADER_FILL
    sheet.cell(row, 2, "Patients").font = HEADER_FONT
    sheet.cell(row, 2).fill = HEADER_FILL
    row += 1
    for key, count in counts:
        sheet.cell(row, 1, key)
        sheet.cell(row, 2, count)
        row += 1
    return row + 1


def _header_row(sheet, headers, row=1):
    for col, text in enumerate(headers, start=1):
        cell = sheet.cell(row, col, text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")


def _autofit(sheet, limit=52):
    for column in sheet.columns:
        width = max((len(str(c.value)) for c in column if c.value is not None), default=8)
        sheet.column_dimensions[get_column_letter(column[0].column)].width = min(width + 3, limit)
