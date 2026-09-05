"""The "not for clinical use" label.

PRD §16.3 lists a missing label as a rollback trigger, so it is enforced by test
rather than left to a reviewer noticing. The label has to survive on the pages a
person reads *and* on the documents that leave the system, because a deck gets
emailed around long after it left the screen that warned about it.
"""

import re
import zipfile
from datetime import date
from io import BytesIO

from django.test import TestCase
from django.urls import reverse

from accounts.models import User
from mdc.models import MDCListing
from patients.models import Patient
from teams.models import MDC, Team

# assertContains is case-sensitive; this is the exact rendered casing.
LABEL = "Not for clinical use"


class PageLabelTests(TestCase):
    def setUp(self):
        self.team = Team.objects.create(consultant="Dr. Test", specialty="Colorectal cancer")
        self.patient = Patient.objects.create(
            name="Test", mrn="990001", date_of_birth=date(1960, 1, 1),
            diagnosis="Low rectal cancer", specialty="COLORECTAL", team=self.team,
        )
        self.user = User.objects.create_user(
            "u1", password="x", role=User.Role.CONSULTANT, team=self.team
        )

    def test_the_sign_in_page_carries_the_label(self):
        # The sign-in page is standalone and does not extend base.html, so it is
        # the one most likely to lose the label unnoticed.
        response = self.client.get(reverse("login"))
        self.assertContains(response, LABEL, status_code=200, msg_prefix="sign-in page")

    def test_every_signed_in_page_carries_the_label(self):
        self.client.force_login(self.user)
        for name, args in [
            ("dashboard", []),
            ("patient_list", []),
            ("patient_detail", [self.patient.pk]),
            ("mdc_board", []),
            ("treatment_list", []),
            ("surgery_schedule", []),
            ("notification_list", []),
            ("team_detail", [self.team.pk]),
        ]:
            with self.subTest(page=name):
                response = self.client.get(reverse(name, args=args))
                self.assertContains(response, LABEL, msg_prefix=name)


class DocumentLabelTests(TestCase):
    """A generated document outlives the page that warned about it."""

    def setUp(self):
        self.team = Team.objects.create(consultant="Dr. Test", specialty="Colorectal cancer")
        self.mdc = MDC.objects.create(name="Gastrointestinal (GI)", meeting_weekday=1)
        self.patient = Patient.objects.create(
            name="Test", mrn="990010", date_of_birth=date(1960, 1, 1),
            diagnosis="Low rectal cancer", specialty="COLORECTAL", team=self.team,
        )
        self.listing = MDCListing.objects.create(
            patient=self.patient, mdc=self.mdc, meeting_date=date(2026, 9, 8)
        )

    def test_every_slide_of_a_deck_carries_the_label(self):
        from mdc import slides

        stream = slides.build_mdc_deck("GI", date(2026, 9, 8), [self.listing])
        archive = zipfile.ZipFile(BytesIO(stream.read()))
        slide_parts = [
            n for n in archive.namelist() if re.match(r"ppt/slides/slide\d+\.xml$", n)
        ]
        self.assertTrue(slide_parts)
        for part in slide_parts:
            with self.subTest(slide=part):
                text = archive.read(part).decode("utf8", "ignore")
                self.assertIn("NOT FOR CLINICAL USE", text)

    def test_both_sheets_of_the_workbook_carry_the_label(self):
        import openpyxl

        from reports.build import build_workbook, collect, period_range

        start, end, label = period_range("month")
        workbook = openpyxl.load_workbook(
            BytesIO(build_workbook(collect(Patient.objects.all(), start, end), label).read())
        )
        for name in workbook.sheetnames:
            with self.subTest(sheet=name):
                sheet = workbook[name]
                found = any(
                    "NOT FOR CLINICAL USE" in str(cell.value or "")
                    for row in sheet.iter_rows(max_row=3)
                    for cell in row
                )
                self.assertTrue(found, f"{name} sheet has no label")

    def test_the_workbook_header_row_survived_the_inserted_label(self):
        """The label shifted the patient sheet down a row; the header must follow."""
        import openpyxl

        from reports.build import build_workbook, collect, period_range

        start, end, label = period_range("month")
        workbook = openpyxl.load_workbook(
            BytesIO(build_workbook(collect(Patient.objects.all(), start, end), label).read())
        )
        sheet = workbook["Patients"]
        self.assertEqual([c.value for c in sheet[2]][:2], ["MRN", "Name"])
        self.assertEqual(sheet["A3"].value, self.patient.mrn)
